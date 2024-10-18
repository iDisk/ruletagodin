import pygame
import random
import math
import time
from openpyxl import load_workbook

# Función para cargar los nombres desde el archivo Excel
def cargar_nombres(archivo_excel):
    wb = load_workbook(archivo_excel)
    hoja = wb.active
    nombres = [row[0] for row in hoja.iter_rows(values_only=True) if row[0] is not None]
    return nombres

# Función para dibujar la ruleta
def dibujar_ruleta(screen, nombres, angulo_offset, ganador=None, fallos=None):
    centro = (300, 300)
    radio = 200
    num_nombres = len(nombres)
    angulo_inicial = angulo_offset

    for i, nombre in enumerate(nombres):
        angulo = angulo_inicial + (i * (360 / num_nombres))
        x = centro[0] + radio * math.cos(math.radians(angulo))
        y = centro[1] + radio * math.sin(math.radians(angulo))
        pos_texto = (x - 20, y - 20)

        if nombre == ganador:
            color = (0, 255, 0)  # Verde para el ganador
        elif nombre in fallos:
            color = (255, 0, 0)  # Rojo para los fallos
        else:
            color = (255, 255, 255)  # Blanco para otros nombres

        texto = font.render(nombre, True, color)
        screen.blit(texto, pos_texto)

# Configuración inicial de pygame
pygame.init()
screen = pygame.display.set_mode((600, 600))  # Dimensión 600x600
pygame.display.set_caption("Ruleta de Participantes")
font = pygame.font.SysFont('Arial', 24)

# Cargar los nombres desde Excel
archivo_excel = "participantes.xlsx"  # Cambia esto a tu archivo
nombres = cargar_nombres(archivo_excel)
participantes_restantes = nombres.copy()
fallos = []

giro_actual = 0
angulo_offset = 0
girando = False
ganador = None
boton_detener_presionado = False
velocidad = 15  # Aumentar la velocidad inicial
reduciendo_velocidad = False

# Lógica del juego
def girar_ruleta_y_determinar():
    global giro_actual, participantes_restantes, fallos, ganador
    ganador = None
    
    if len(participantes_restantes) < 3:
        return random.choice(participantes_restantes)  # Si quedan pocos, elige al azar

    giro_actual += 1

    if giro_actual % 3 == 0:
        # Tercer giro, elige ganador
        ganador = random.choice(participantes_restantes)
        participantes_restantes.remove(ganador)
    else:
        # Primeros dos giros, elimina un participante
        perdedor = random.choice(participantes_restantes)
        fallos.append(perdedor)
        participantes_restantes.remove(perdedor)

    return ganador

# Dibujar el botón
def dibujar_boton(screen, texto, pos):
    font_boton = pygame.font.SysFont('Arial', 26)
    texto_superficie = font_boton.render(texto, True, (255, 255, 255))
    rect = texto_superficie.get_rect(center=pos)
    pygame.draw.rect(screen, (0, 100, 255), rect.inflate(20, 20), border_radius=10)
    screen.blit(texto_superficie, rect)
    return rect

# Bucle principal
corriendo = True
while corriendo:
    screen.fill((30, 30, 30))  # Fondo gris oscuro
    
    # Dibuja la ruleta
    dibujar_ruleta(screen, participantes_restantes, angulo_offset, ganador, fallos)
    
    # Dibuja el botón de detener
    rect_boton = dibujar_boton(screen, "Detener", (300, 550))
    
    # Procesa los eventos
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            corriendo = False
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_SPACE:  # Presiona espacio para empezar a girar
                girando = True
                boton_detener_presionado = False
                reduciendo_velocidad = False
                velocidad = 15  # Velocidad más rápida al inicio
        if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
            if rect_boton.collidepoint(event.pos):
                boton_detener_presionado = True

    if girando and not boton_detener_presionado:
        angulo_offset += velocidad  # Ruleta rápida
    elif boton_detener_presionado and girando:
        reduciendo_velocidad = True  # Comenzar a reducir la velocidad
    
    if reduciendo_velocidad:
        velocidad = max(1, velocidad - 0.2)  # Reduce la velocidad gradualmente
        angulo_offset += velocidad
        if velocidad <= 1:  # Cuando la velocidad es mínima, detener el giro
            ganador = girar_ruleta_y_determinar()
            girando = False
            reduciendo_velocidad = False

    pygame.display.flip()
    pygame.time.wait(30)  # Controla la velocidad del ciclo de actualización

pygame.quit()