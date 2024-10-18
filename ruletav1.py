import tkinter as tk
import random
import time
from openpyxl import load_workbook

# Función para cargar los nombres desde el archivo Excel
def cargar_nombres(archivo_excel):
    wb = load_workbook(archivo_excel)
    hoja = wb.active
    nombres = []
    for row in hoja.iter_rows(values_only=True):
        nombres.append(row[0])
    return nombres

# Función para simular el giro de la ruleta
def girar_ruleta():
    global giro_actual, nombres

    resultado_label.config(text="Girando...")
    root.update()

    # Simulación de giro lento
    for _ in range(30):
        nombre_aleatorio = random.choice(nombres)
        resultado_label.config(text=nombre_aleatorio)
        root.update()
        time.sleep(0.1)

    # Lógica de ganadores
    giro_actual += 1
    if giro_actual % 3 == 0:
        ganador = random.choice(nombres)
        resultado_label.config(text=f"Ganador: {ganador}", fg="green")
    else:
        resultado_label.config(text="¡Fallaste! Intenta de nuevo.", fg="red")

    root.update()

# Cargar nombres desde Excel
archivo_excel = "participantes.xlsx"  # Cambia a tu archivo Excel
nombres = cargar_nombres(archivo_excel)

# Crear la interfaz gráfica
root = tk.Tk()
root.title("Ruleta de Participantes")

giro_actual = 0

# Etiqueta para mostrar el resultado
resultado_label = tk.Label(root, text="Presiona para girar", font=("Helvetica", 24))
resultado_label.pack(pady=20)

# Botón para girar la ruleta
girar_button = tk.Button(root, text="Girar", font=("Helvetica", 18), command=girar_ruleta)
girar_button.pack(pady=20)

# Iniciar la aplicación
root.mainloop()
